# views/api_views.py
from django.http import JsonResponse
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.cache import cache
from django.utils import timezone

from django.db import transaction
from celery.result import AsyncResult

from ..models import Simulation, ResultSimulation, SimulationShareToken
from ..services.simulation_service import SimulationService
from ..tasks import execute_simulation_async

import logging

logger = logging.getLogger(__name__)

# Celery task states that mean the job is still running
_ACTIVE_STATES = {'PENDING', 'RECEIVED', 'STARTED', 'PROGRESS', 'RETRY'}


def _observed_series(results, key):
    """Serie con los períodos que SÍ registraron la variable.

    Rellenar los ausentes con 0 convertía la falta de datos en ceros observados:
    sumas deprimidas, probabilidad de pérdida artificialmente baja y una corrida
    vacía compitiendo por ser la mejor.
    """
    series = []
    for row in results:
        variables = getattr(row, 'variables', None) or {}
        value = variables.get(key)
        if value is None:
            continue
        try:
            series.append(float(value))
        except (TypeError, ValueError):
            continue
    return series


class SimulationProgressView(LoginRequiredMixin, View):
    """Check simulation progress — prefers live Celery task state, falls back to DB."""

    def get(self, request, simulation_id):
        try:
            simulation = Simulation.objects.get(
                id=simulation_id,
                fk_questionary_result__fk_questionary__fk_product__fk_business__fk_user=request.user
            )

            # ── 1. Real-time Celery progress ────────────────────────────────
            task_id = cache.get(f"sim_task_id:{simulation_id}")
            if task_id:
                result = AsyncResult(task_id)
                task_state = result.state
                meta = result.info if isinstance(result.info, dict) else {}

                if task_state == 'SUCCESS':
                    return JsonResponse({
                        'simulation_id': simulation_id,
                        'is_completed': True,
                        'status': 'completed',
                        'progress': 100,
                        'message': 'Simulación completada',
                        'phase': 'done',
                        'task_id': task_id,
                    })

                if task_state == 'FAILURE':
                    return JsonResponse({
                        'simulation_id': simulation_id,
                        'is_completed': False,
                        'status': 'failed',
                        'progress': 0,
                        'message': str(result.result) if result.result else 'Error en la simulación',
                        'phase': 'error',
                        'task_id': task_id,
                    })

                if task_state in _ACTIVE_STATES:
                    current = meta.get('current', 0)
                    total = meta.get('total', 100)
                    pct = round(current / total * 100, 2) if total else 0
                    return JsonResponse({
                        'simulation_id': simulation_id,
                        'is_completed': False,
                        'status': 'processing',
                        'progress': pct,
                        'message': meta.get('status', 'Procesando...'),
                        'phase': meta.get('phase', 'simulation'),
                        'period': meta.get('period'),
                        'total_periods': meta.get('total_periods'),
                        'task_id': task_id,
                    })

            # ── 2. Fallback: derive progress from persisted DB results ───────
            total_days = simulation.quantity_time
            completed_days = simulation.results.filter(is_active=True).count()
            progress = round(completed_days / total_days * 100, 2) if total_days > 0 else 0

            if simulation.is_completed:
                status = 'completed'
                message = 'Simulación completada'
            elif completed_days == 0:
                status = 'pending'
                message = 'Simulación pendiente de ejecución'
            else:
                status = 'processing'
                message = f'Procesando día {completed_days} de {total_days}'

            return JsonResponse({
                'simulation_id': simulation_id,
                'is_completed': simulation.is_completed,
                'status': status,
                'progress': progress,
                'message': message,
                'completed_days': completed_days,
                'total_days': total_days,
            })

        except Simulation.DoesNotExist:
            return JsonResponse({'error': 'Simulación no encontrada'}, status=404)
        except Exception as e:
            logger.error("Error checking progress for simulation %s: %s", simulation_id, e)
            return JsonResponse({'error': str(e)}, status=500)

class SimulationRetryView(LoginRequiredMixin, View):
    """Retry failed simulation"""
    
    def post(self, request, simulation_id):
        try:
            simulation = Simulation.objects.get(
                id=simulation_id,
                fk_questionary_result__fk_questionary__fk_product__fk_business__fk_user=request.user
            )
            
            with transaction.atomic():
                simulation.results.all().delete()
                simulation.demands.all().delete()
                simulation.is_completed = False
                simulation.save(update_fields=['is_completed'])

            sim_id = simulation.id
            transaction.on_commit(lambda: execute_simulation_async.delay(sim_id))

            return JsonResponse({
                'success': True,
                'message': 'Simulación reiniciada. Procesando en segundo plano...'
            })
            
        except Exception as e:
            logger.error(f"Error retrying simulation: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)

class SimulationDuplicateView(LoginRequiredMixin, View):
    """Duplicate simulation"""
    
    def post(self, request, simulation_id):
        try:
            original = Simulation.objects.get(
                id=simulation_id,
                fk_questionary_result__fk_questionary__fk_product__fk_business__fk_user=request.user
            )
            
            # Create duplicate
            duplicate = Simulation.objects.create(
                fk_questionary_result=original.fk_questionary_result,
                quantity_time=original.quantity_time,
                unit_time=original.unit_time,
                demand_history=original.demand_history,
                fk_fdp=original.fk_fdp,
                confidence_level=original.confidence_level,
                random_seed=original.random_seed,
                is_active=True
            )
            
            return JsonResponse({
                'success': True,
                'new_id': duplicate.id,
                'message': 'Simulación duplicada correctamente'
            })
            
        except Exception as e:
            logger.error(f"Error duplicating simulation: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)

class SimulationStartView(LoginRequiredMixin, View):
    """Start pending simulation"""
    
    def post(self, request, simulation_id):
        try:
            simulation = Simulation.objects.get(
                id=simulation_id,
                fk_questionary_result__fk_questionary__fk_product__fk_business__fk_user=request.user
            )
            
            # Check if already has results
            if simulation.results.exists():
                return JsonResponse({
                    'success': False,
                    'message': 'La simulación ya fue ejecutada. Use reintentar para volver a ejecutar.'
                })
            
            task = execute_simulation_async.delay(simulation.id)

            return JsonResponse({
                'success': True,
                'task_id': task.id,
                'message': 'Simulación iniciada. Procesando en segundo plano...'
            })
            
        except Exception as e:
            logger.error(f"Error starting simulation: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)

class SimulationDeleteView(LoginRequiredMixin, View):
    """Delete simulation (soft delete)"""
    
    def post(self, request, simulation_id):
        try:
            simulation = Simulation.objects.get(
                id=simulation_id,
                fk_questionary_result__fk_questionary__fk_product__fk_business__fk_user=request.user
            )
            
            # Soft delete
            simulation.is_active = False
            simulation.save()
            
            # Also deactivate related records
            simulation.results.update(is_active=False)
            simulation.demands.update(is_active=False)
            
            return JsonResponse({
                'success': True,
                'message': 'Simulación eliminada correctamente'
            })
            
        except Exception as e:
            logger.error(f"Error deleting simulation: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)


# ── FASE 4.1 — Scenario comparison ──────────────────────────────────────────

class SimulationCompareView(LoginRequiredMixin, View):
    """
    Compare up to 4 simulations side-by-side.
    GET /api/compare/?ids=1,2,3
    Returns JSON with per-simulation KPIs ready for a comparison table/chart.
    """

    def get(self, request):
        raw = request.GET.get('ids', '')
        try:
            ids = [int(x) for x in raw.split(',') if x.strip()][:4]
        except ValueError:
            return JsonResponse({'error': 'ids must be comma-separated integers'}, status=400)

        if len(ids) < 2:
            return JsonResponse({'error': 'Provide at least 2 simulation ids'}, status=400)

        user_qs = Simulation.objects.filter(
            id__in=ids,
            fk_questionary_result__fk_questionary__fk_product__fk_business__fk_user=request.user,
        ).select_related(
            'fk_questionary_result__fk_questionary__fk_product',
            'fk_fdp',
        ).prefetch_related('results')

        found = {s.id: s for s in user_qs}
        missing = set(ids) - set(found)
        if missing:
            return JsonResponse({'error': f'Simulations not found or access denied: {missing}'}, status=404)

        import numpy as np

        comparisons = []
        for sid in ids:
            sim = found[sid]
            results = list(sim.results.filter(is_active=True).order_by('date'))
            if not results:
                comparisons.append({'simulation_id': sid, 'no_results': True})
                continue

            # Sólo períodos con la variable observada. Con `.get(clave, 0)` una
            # corrida sin variables financieras entraba a la comparación como un
            # negocio que factura 0, no pierde nunca (0 no es < 0) y competía por
            # ser "ganador".
            profit_series = _observed_series(results, 'GT')
            revenue_series = _observed_series(results, 'IT')
            cost_series = _observed_series(results, 'TG')
            demand_series = [float(r.demand_mean) for r in results]

            if not profit_series:
                comparisons.append({
                    'simulation_id': sid,
                    'product_name': sim.fk_questionary_result.fk_questionary.fk_product.name,
                    'duration_days': sim.duration_in_days,
                    'status': 'sin_datos_financieros',
                    'observed_periods': 0,
                    'total_periods': len(results),
                })
                continue

            arr = np.array(profit_series)

            total_profit = float(arr.sum())
            total_revenue = float(np.array(revenue_series).sum()) if revenue_series else None
            # Sin ingresos observados el margen es indefinido, no 0%.
            profit_margin = (round(total_profit / total_revenue * 100, 2)
                             if total_revenue else None)

            var_95 = float(np.percentile(arr, 5))
            prob_loss = float(np.mean(arr < 0)) * 100

            comparisons.append({
                'simulation_id': sid,
                'product_name': sim.fk_questionary_result.fk_questionary.fk_product.name,
                'duration_days': sim.duration_in_days,
                'distribution': sim.fk_fdp.name if sim.fk_fdp else '—',
                'confidence_level': sim.confidence_level,
                'date_created': sim.date_created.strftime('%d/%m/%Y') if sim.date_created else '—',
                'status': 'ok',
                'observed_periods': len(profit_series),
                'total_periods': len(results),
                'total_profit': round(total_profit, 2),
                'total_revenue': round(total_revenue, 2) if total_revenue is not None else None,
                'total_costs': (round(float(np.array(cost_series).sum()), 2)
                                if cost_series else None),
                'profit_margin_pct': profit_margin,
                'mean_daily_profit': round(float(arr.mean()), 2),
                'std_daily_profit': round(float(arr.std()), 2),
                'var_95': round(var_95, 2),
                'probability_loss_pct': round(prob_loss, 1),
                'mean_demand': round(float(np.mean(demand_series)), 2),
                'profitable': total_profit > 0,
            })

        # Determine winner (highest total_profit among completed)
        completed = [c for c in comparisons if 'total_profit' in c]
        if completed:
            winner_id = max(completed, key=lambda c: c['total_profit'])['simulation_id']
        else:
            winner_id = None

        return JsonResponse({
            'success': True,
            'comparisons': comparisons,
            'winner_simulation_id': winner_id,
            'metrics_explained': {
                'total_profit': 'Suma de GT (ganancia) en todos los períodos',
                'var_95': 'Peor ganancia esperada en el 5% de días (VaR 95%)',
                'probability_loss_pct': 'Porcentaje de días con ganancia negativa',
            }
        })


# ── FASE 4.2 — Sensitivity analysis (correlation-based tornado) ───────────

class SimulationSensitivityView(LoginRequiredMixin, View):
    """
    Compute input-variable sensitivity against profit (GT) using Pearson
    correlation over the period time-series. No re-run needed.
    GET /api/sensitivity/<simulation_id>/
    Returns entries sorted by |correlation| descending — ready for a tornado chart.
    """

    def get(self, request, simulation_id):
        try:
            simulation = Simulation.objects.get(
                id=simulation_id,
                fk_questionary_result__fk_questionary__fk_product__fk_business__fk_user=request.user,
            )
        except Simulation.DoesNotExist:
            return JsonResponse({'error': 'Simulación no encontrada'}, status=404)

        results = list(
            simulation.results.filter(is_active=True).order_by('date')
        )
        if len(results) < 4:
            return JsonResponse({'error': 'Se necesitan al menos 4 períodos de resultados'}, status=400)

        import numpy as np

        profit = np.array([float(r.variables.get('GT', 0)) for r in results])
        if profit.std() == 0:
            return JsonResponse({'error': 'No hay variación en la ganancia GT'}, status=400)

        # Collect all variable names present
        all_keys: set = set()
        for r in results:
            all_keys.update(r.variables.keys())
        all_keys.discard('GT')

        entries = []
        for key in sorted(all_keys):
            series = np.array([float(r.variables.get(key, 0)) for r in results])
            if series.std() == 0:
                continue
            corr = float(np.corrcoef(series, profit)[0, 1])
            if np.isnan(corr):
                continue
            entries.append({
                'variable': key,
                'correlation': round(corr, 4),
                'abs_correlation': round(abs(corr), 4),
                'direction': 'positive' if corr > 0 else 'negative',
                'impact_label': (
                    'Alto' if abs(corr) >= 0.6
                    else 'Medio' if abs(corr) >= 0.3
                    else 'Bajo'
                ),
            })

        entries.sort(key=lambda e: e['abs_correlation'], reverse=True)

        return JsonResponse({
            'success': True,
            'simulation_id': simulation_id,
            'periods_analyzed': len(results),
            'sensitivity': entries[:20],  # top 20 drivers
            'chart_hint': 'Render as horizontal bar chart sorted by abs_correlation. Positive=green, negative=red.',
        })


# ── FASE 4.3 — Simulation history & trends ────────────────────────────────

class SimulationHistoryView(LoginRequiredMixin, View):
    """
    Aggregated simulation history for a business, grouped by month.
    GET /api/history/?business_id=<id>
    """

    def get(self, request):
        from django.db.models import Count, Avg
        from collections import defaultdict

        business_id = request.GET.get('business_id') or request.session.get('business_id')
        if not business_id:
            return JsonResponse({'error': 'business_id requerido'}, status=400)

        try:
            business_id = int(business_id)
        except (ValueError, TypeError):
            return JsonResponse({'error': 'business_id inválido'}, status=400)

        from business.models import Business
        try:
            business = Business.objects.get(id=business_id, fk_user=request.user)
        except Business.DoesNotExist:
            return JsonResponse({'error': 'Negocio no encontrado'}, status=404)

        sims = (
            Simulation.objects.filter(
                fk_questionary_result__fk_questionary__fk_product__fk_business=business,
                is_active=True,
            )
            .select_related('fk_questionary_result__fk_questionary__fk_product')
            .order_by('date_created')
        )

        monthly: dict = defaultdict(lambda: {'count': 0, 'completed': 0, 'total_days': 0})
        product_counts: dict = defaultdict(int)

        import numpy as np

        for sim in sims:
            month_key = sim.date_created.strftime('%b %Y') if sim.date_created else 'Unknown'
            monthly[month_key]['count'] += 1
            if sim.is_completed:
                monthly[month_key]['completed'] += 1
            monthly[month_key]['total_days'] += sim.duration_in_days
            product_counts[sim.fk_questionary_result.fk_questionary.fk_product.name] += 1

        timeline = [
            {
                'month': k,
                'total': v['count'],
                'completed': v['completed'],
                'avg_days': round(v['total_days'] / v['count'], 1) if v['count'] else 0,
            }
            for k, v in monthly.items()
        ]

        top_products = sorted(product_counts.items(), key=lambda x: x[1], reverse=True)[:5]

        return JsonResponse({
            'success': True,
            'business_name': business.name,
            'total_simulations': sims.count(),
            'completed_simulations': sims.filter(is_completed=True).count(),
            'timeline': timeline,
            'top_products': [{'name': n, 'count': c} for n, c in top_products],
        })


# ── FASE 4.4 — Enriched export (Excel multi-sheet + JSON) ─────────────────

class SimulationExportView(LoginRequiredMixin, View):
    """
    Export simulation results in enriched formats.
    GET /api/export/<simulation_id>/?format=excel|json
    """

    def get(self, request, simulation_id):
        fmt = request.GET.get('format', 'json').lower()

        try:
            simulation = Simulation.objects.select_related(
                'fk_questionary_result__fk_questionary__fk_product__fk_business',
                'fk_fdp',
            ).get(
                id=simulation_id,
                fk_questionary_result__fk_questionary__fk_product__fk_business__fk_user=request.user,
            )
        except Simulation.DoesNotExist:
            return JsonResponse({'error': 'Simulación no encontrada'}, status=404)

        results = list(simulation.results.filter(is_active=True).order_by('date'))

        if fmt == 'excel':
            return self._export_excel(simulation, results)
        return self._export_json(simulation, results)

    # ── Excel ────────────────────────────────────────────────────────────────

    def _export_excel(self, simulation, results):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            from django.http import HttpResponse
            import io, numpy as np

            wb = openpyxl.Workbook()

            # ── Sheet 1: Summary ─────────────────────────────────────────────
            ws_sum = wb.active
            ws_sum.title = 'Resumen Ejecutivo'
            product = simulation.fk_questionary_result.fk_questionary.fk_product

            header_fill = PatternFill('solid', fgColor='1d4ed8')
            header_font = Font(color='FFFFFF', bold=True)

            ws_sum.append(['Simulación FindemproAI — Resumen Ejecutivo'])
            ws_sum['A1'].font = Font(bold=True, size=14)
            ws_sum.append([])
            ws_sum.append(['Campo', 'Valor'])
            for cell in ws_sum[3]:
                cell.fill = header_fill; cell.font = header_font

            info = [
                ('ID Simulación', simulation.id),
                ('Producto', product.name),
                ('Negocio', product.fk_business.name),
                ('Duración', f'{simulation.quantity_time} {simulation.unit_time}'),
                ('Distribución FDP', simulation.fk_fdp.name if simulation.fk_fdp else '—'),
                ('Confianza', f'{simulation.confidence_level * 100:.0f}%'),
                ('Fecha creación', simulation.date_created.strftime('%d/%m/%Y %H:%M') if simulation.date_created else '—'),
                ('Períodos simulados', len(results)),
            ]

            if results:
                profit = np.array([float(r.variables.get('GT', 0)) for r in results])
                revenue = np.array([float(r.variables.get('IT', 0)) for r in results])
                costs = np.array([float(r.variables.get('TG', 0)) for r in results])
                info += [
                    ('', ''),
                    ('Ganancia Total (GT)', round(float(profit.sum()), 2)),
                    ('Ingreso Total (IT)', round(float(revenue.sum()), 2)),
                    ('Costo Total (TG)', round(float(costs.sum()), 2)),
                    ('Margen (%)', round(float(profit.sum() / revenue.sum() * 100), 2) if revenue.sum() else 0),
                    ('Ganancia Media/Período', round(float(profit.mean()), 2)),
                    ('Desv. Estándar (σ)', round(float(profit.std()), 2)),
                    ('VaR 95%', round(float(np.percentile(profit, 5)), 2)),
                    ('CVaR 95%', round(float(profit[profit <= np.percentile(profit, 5)].mean()), 2) if any(profit <= np.percentile(profit, 5)) else 0),
                    ('Prob. Pérdida (%)', round(float(np.mean(profit < 0) * 100), 1)),
                ]

            for row in info:
                ws_sum.append(list(row))

            ws_sum.column_dimensions['A'].width = 28
            ws_sum.column_dimensions['B'].width = 22

            # ── Sheet 2: Daily Results ────────────────────────────────────────
            ws_daily = wb.create_sheet('Resultados Diarios')
            if results:
                all_vars = sorted({k for r in results for k in r.variables})
                headers = ['Período', 'Fecha', 'Demanda Media'] + all_vars
                ws_daily.append(headers)
                for cell in ws_daily[1]:
                    cell.fill = header_fill; cell.font = header_font

                for i, r in enumerate(results, 1):
                    row = [i, r.date.strftime('%d/%m/%Y'), float(r.demand_mean)]
                    row += [r.variables.get(k, '') for k in all_vars]
                    ws_daily.append(row)

                for col_idx in range(1, len(headers) + 1):
                    ws_daily.column_dimensions[get_column_letter(col_idx)].width = 14

            # ── Sheet 3: Risk Metrics ─────────────────────────────────────────
            ws_risk = wb.create_sheet('Métricas de Riesgo')
            ws_risk.append(['Percentil', 'GT (Bs.)'])
            for cell in ws_risk[1]:
                cell.fill = header_fill; cell.font = header_font
            if results:
                for pct in [5, 10, 25, 50, 75, 90, 95]:
                    ws_risk.append([f'P{pct}', round(float(np.percentile(profit, pct)), 2)])
            ws_risk.column_dimensions['A'].width = 14
            ws_risk.column_dimensions['B'].width = 16

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            filename = f"simulacion_{simulation.id}_{product.name.replace(' ', '_')}.xlsx"
            response = HttpResponse(
                buf.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except ImportError:
            return JsonResponse({'error': 'openpyxl no está instalado. pip install openpyxl'}, status=500)
        except Exception as exc:
            logger.error("Excel export error for simulation %s: %s", simulation.id, exc)
            return JsonResponse({'error': str(exc)}, status=500)

    # ── JSON ─────────────────────────────────────────────────────────────────

    def _export_json(self, simulation, results):
        import numpy as np
        from django.http import HttpResponse
        import json as _json

        product = simulation.fk_questionary_result.fk_questionary.fk_product
        profit = np.array([float(r.variables.get('GT', 0)) for r in results]) if results else np.array([])

        payload = {
            'simulation_id': simulation.id,
            'product': product.name,
            'business': product.fk_business.name,
            'duration_days': simulation.duration_in_days,
            'fdp': simulation.fk_fdp.name if simulation.fk_fdp else None,
            'confidence_level': simulation.confidence_level,
            'date_created': simulation.date_created.isoformat() if simulation.date_created else None,
            'risk_summary': {
                'total_profit': round(float(profit.sum()), 2),
                'mean': round(float(profit.mean()), 2),
                'std': round(float(profit.std()), 2),
                'var_95': round(float(np.percentile(profit, 5)), 2),
                'probability_loss_pct': round(float(np.mean(profit < 0) * 100), 1),
            } if profit.size else {},
            'periods': [
                {
                    'period': i + 1,
                    'date': r.date.isoformat(),
                    'demand_mean': float(r.demand_mean),
                    'variables': r.variables,
                }
                for i, r in enumerate(results)
            ],
        }

        filename = f"simulacion_{simulation.id}.json"
        response = HttpResponse(
            _json.dumps(payload, ensure_ascii=False, indent=2),
            content_type='application/json',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ── Share by temporary link ──────────────────────────────────────────────────

class SimulationShareCreateView(LoginRequiredMixin, View):
    """
    POST /api/share/<simulation_id>/
    Creates a 7-day share token. Returns the public URL.
    Body (optional JSON): { "days": 7 }
    """

    def post(self, request, simulation_id):
        try:
            simulation = Simulation.objects.get(
                id=simulation_id,
                fk_questionary_result__fk_questionary__fk_product__fk_business__fk_user=request.user,
            )
        except Simulation.DoesNotExist:
            return JsonResponse({'error': 'Simulación no encontrada'}, status=404)

        try:
            import json as _json
            body = _json.loads(request.body or '{}')
            days = min(int(body.get('days', 7)), 30)
        except (ValueError, KeyError):
            days = 7

        share = SimulationShareToken.create_for(simulation, request.user, days=days)
        share_url = request.build_absolute_uri(f'/simulate/share/{share.token}/')

        return JsonResponse({
            'token':      str(share.token),
            'url':        share_url,
            'expires_at': share.expires_at.isoformat(),
            'days':       days,
        }, status=201)


class SimulationShareAccessView(View):
    """
    GET /simulate/share/<token>/
    Public (no auth) access to a shared simulation. Returns JSON summary.
    """

    def get(self, request, token):
        try:
            share = SimulationShareToken.objects.select_related('simulation').get(token=token)
        except SimulationShareToken.DoesNotExist:
            return JsonResponse({'error': 'Enlace no encontrado'}, status=404)

        if not share.is_valid():
            return JsonResponse({'error': 'El enlace ha expirado o fue desactivado'}, status=410)

        share.access_count += 1
        share.save(update_fields=['access_count'])

        sim = share.simulation
        results = list(sim.results.filter(is_active=True).order_by('date').values(
            'date', 'demand_mean', 'variables',
        ))

        return JsonResponse({
            'simulation_id': sim.id,
            'product':       str(sim.fk_questionary_result),
            'duration_days': sim.duration_in_days,
            'is_completed':  sim.is_completed,
            'expires_at':    share.expires_at.isoformat(),
            'access_count':  share.access_count,
            'results':       [
                {
                    'date': str(r['date']),
                    'demand_mean': float(r['demand_mean']),
                    'profit': float((r['variables'] or {}).get('GT', 0)),
                    'revenue': float((r['variables'] or {}).get('IT', 0)),
                }
                for r in results
            ],
        })